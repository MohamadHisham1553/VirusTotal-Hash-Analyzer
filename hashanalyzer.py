"""
VirusTotal Hash Analyzer
========================
Queries VirusTotal API for file hashes and generates a detailed Excel report
with conditional formatting and charts.

Requirements:
    pip install requests openpyxl

Usage:
    Run the script and enter hashes when prompted (one per line, blank line to finish),
    or pass them directly via the HASHES list at the bottom of this file.
"""

import os
import sys
import time
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

# ─────────────────────────── CONFIGURATION ────────────────────────────────── #

API_KEY   = "a5e705584c31e9b786062263fe6aa71b26fff3c3fa04767444317a2797a42301"
BASE_URL  = "https://www.virustotal.com/api/v3/files/{}"
DELAY     = 15          # seconds between requests (free tier: 4 req/min)
OUTPUT    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VirusTotal_Report.xlsx")

# ─────────────────────────── COLOR PALETTE ────────────────────────────────── #

CLR = {
    "header_bg"    : "1F3864",   # dark navy
    "header_fg"    : "FFFFFF",
    "title_bg"     : "2E75B6",   # accent blue
    "subheader_bg" : "D6E4F7",
    "clean_bg"     : "C6EFCE",   # green
    "clean_fg"     : "276221",
    "malicious_bg" : "FFC7CE",   # red
    "malicious_fg" : "9C0006",
    "suspicious_bg": "FFEB9C",   # yellow
    "suspicious_fg": "9C5700",
    "unknown_bg"   : "F2F2F2",
    "unknown_fg"   : "595959",
    "row_alt"      : "EBF3FB",
    "border"       : "B8CCE4",
}

def clr(hex_code):
    return PatternFill("solid", start_color=hex_code, end_color=hex_code)

def verdict_style(malicious, suspicious, total_engines):
    """Return (bg, fg, label) based on detection counts."""
    if total_engines == 0:
        return CLR["unknown_bg"], CLR["unknown_fg"], "Unknown"
    ratio = malicious / total_engines
    if malicious == 0 and suspicious == 0:
        return CLR["clean_bg"], CLR["clean_fg"], "Clean"
    if malicious > 0 or ratio >= 0.05:
        return CLR["malicious_bg"], CLR["malicious_fg"], "Malicious"
    return CLR["suspicious_bg"], CLR["suspicious_fg"], "Suspicious"


# ─────────────────────────── API FUNCTIONS ────────────────────────────────── #

def query_hash(file_hash: str) -> dict:
    """
    Query VirusTotal v3 API for a single file hash.
    Returns a flat dict with all key fields, or an error dict.
    """
    headers = {"x-apikey": API_KEY, "Accept": "application/json"}
    url     = BASE_URL.format(file_hash.strip())

    try:
        resp = requests.get(url, headers=headers, timeout=30)
    except requests.RequestException as exc:
        return {"hash": file_hash, "error": str(exc)}

    if resp.status_code == 404:
        return {"hash": file_hash, "error": "Not found in VirusTotal"}
    if resp.status_code == 429:
        return {"hash": file_hash, "error": "Rate limit exceeded – wait and retry"}
    if resp.status_code != 200:
        return {"hash": file_hash, "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}

    data  = resp.json().get("data", {})
    attrs = data.get("attributes", {})

    stats = attrs.get("last_analysis_stats", {})
    malicious   = stats.get("malicious", 0)
    suspicious  = stats.get("suspicious", 0)
    undetected  = stats.get("undetected", 0)
    harmless    = stats.get("harmless", 0)
    timeout_c   = stats.get("timeout", 0)
    unsupported = stats.get("type-unsupported", 0)
    failure     = stats.get("failure", 0)
    total_eng   = malicious + suspicious + undetected + harmless + timeout_c + unsupported + failure

    # First-seen / last-seen timestamps
    def ts(epoch):
        return datetime.utcfromtimestamp(epoch).strftime("%Y-%m-%d %H:%M UTC") if epoch else "N/A"

    # Collect popular AV engine verdicts (up to 10)
    results   = attrs.get("last_analysis_results", {})
    detections = [
        f"{eng}: {res.get('result', 'N/A')}"
        for eng, res in results.items()
        if res.get("category") in ("malicious", "suspicious")
    ][:10]

    names = attrs.get("meaningful_name") or attrs.get("names", ["N/A"])
    if isinstance(names, list):
        names = ", ".join(names[:3])

    return {
        "hash"            : file_hash.strip(),
        "error"           : None,
        "name"            : names,
        "file_type"       : attrs.get("type_description", attrs.get("type_tag", "N/A")),
        "file_size"       : attrs.get("size", "N/A"),
        "md5"             : attrs.get("md5", "N/A"),
        "sha1"            : attrs.get("sha1", "N/A"),
        "sha256"          : attrs.get("sha256", "N/A"),
        "malicious"       : malicious,
        "suspicious"      : suspicious,
        "undetected"      : undetected,
        "harmless"        : harmless,
        "total_engines"   : total_eng,
        "detection_ratio" : f"{malicious}/{total_eng}" if total_eng else "N/A",
        "reputation"      : attrs.get("reputation", "N/A"),
        "community_votes_harmless"  : attrs.get("total_votes", {}).get("harmless", 0),
        "community_votes_malicious" : attrs.get("total_votes", {}).get("malicious", 0),
        "first_seen"      : ts(attrs.get("first_submission_date")),
        "last_seen"       : ts(attrs.get("last_submission_date")),
        "last_analysis"   : ts(attrs.get("last_analysis_date")),
        "times_submitted" : attrs.get("times_submitted", "N/A"),
        "tags"            : ", ".join(attrs.get("tags", [])) or "N/A",
        "detections_list" : "\n".join(detections) if detections else "None",
        "ssdeep"          : attrs.get("ssdeep", "N/A"),
        "tlsh"            : attrs.get("tlsh", "N/A"),
    }


def query_all(hashes: list) -> list:
    """Query all hashes with rate-limiting and progress output."""
    results = []
    total   = len(hashes)
    for idx, h in enumerate(hashes, 1):
        print(f"  [{idx}/{total}] Querying: {h.strip()} ...", end=" ", flush=True)
        result = query_hash(h)
        if result.get("error"):
            print(f"ERROR – {result['error']}")
        else:
            print(f"OK  ({result['malicious']} malicious / {result['total_engines']} engines)")
        results.append(result)
        if idx < total:
            print(f"    Waiting {DELAY}s (API rate limit) ...", flush=True)
            time.sleep(DELAY)
    return results


# ─────────────────────────── EXCEL REPORT ─────────────────────────────────── #

COLUMNS = [
    ("Hash (Input)",          28),
    ("Verdict",               14),
    ("Detection Ratio",       16),
    ("Malicious",             12),
    ("Suspicious",            12),
    ("Undetected",            12),
    ("Total Engines",         14),
    ("File Name",             28),
    ("File Type",             18),
    ("File Size (bytes)",     18),
    ("Reputation",            12),
    ("Community: Harmless",   20),
    ("Community: Malicious",  20),
    ("First Seen (UTC)",      22),
    ("Last Seen (UTC)",       22),
    ("Last Analysis (UTC)",   22),
    ("Times Submitted",       16),
    ("Tags",                  22),
    ("MD5",                   34),
    ("SHA1",                  42),
    ("SHA256",                66),
    ("Top Detections",        50),
    ("Error",                 40),
]

def thin_border(sides=("left","right","top","bottom")):
    s = Side(style="thin", color=CLR["border"])
    kw = {k: s for k in sides}
    return Border(**kw)

def build_excel(results: list, output_path: str):
    wb = Workbook()

    # ── Sheet 1: Summary Report ──────────────────────────────────────────── #
    ws = wb.active
    ws.title = "Hash Report"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"VirusTotal Hash Analysis Report  –  Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    title_cell.font      = Font(name="Arial", size=14, bold=True, color=CLR["header_fg"])
    title_cell.fill      = clr(CLR["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, (col_name, col_w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", size=10, bold=True, color=CLR["header_fg"])
        cell.fill      = clr(CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[2].height = 32

    # Data rows
    for row_idx, rec in enumerate(results, 3):
        is_alt = (row_idx % 2 == 0)
        mal    = rec.get("malicious", 0) or 0
        sus    = rec.get("suspicious", 0) or 0
        tot    = rec.get("total_engines", 0) or 0
        bg_v, fg_v, label = verdict_style(mal, sus, tot if not rec.get("error") else 0)

        row_data = [
            rec.get("hash", ""),
            label if not rec.get("error") else "Error",
            rec.get("detection_ratio", "N/A"),
            mal,
            sus,
            rec.get("undetected", "N/A"),
            tot,
            rec.get("name", "N/A"),
            rec.get("file_type", "N/A"),
            rec.get("file_size", "N/A"),
            rec.get("reputation", "N/A"),
            rec.get("community_votes_harmless", "N/A"),
            rec.get("community_votes_malicious", "N/A"),
            rec.get("first_seen", "N/A"),
            rec.get("last_seen", "N/A"),
            rec.get("last_analysis", "N/A"),
            rec.get("times_submitted", "N/A"),
            rec.get("tags", "N/A"),
            rec.get("md5", "N/A"),
            rec.get("sha1", "N/A"),
            rec.get("sha256", "N/A"),
            rec.get("detections_list", "N/A"),
            rec.get("error") or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            cell.border    = thin_border()

            # Alternating row background
            base_fill = CLR["row_alt"] if is_alt else "FFFFFF"
            cell.fill  = clr(base_fill)

        # Verdict cell coloring (column 2)
        v_cell       = ws.cell(row=row_idx, column=2)
        v_cell.fill  = clr(bg_v)
        v_cell.font  = Font(name="Arial", size=9, bold=True, color=fg_v)
        v_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Malicious count cell – color if > 0
        m_cell = ws.cell(row=row_idx, column=4)
        if isinstance(mal, int) and mal > 0:
            m_cell.fill = clr(CLR["malicious_bg"])
            m_cell.font = Font(name="Arial", size=9, bold=True, color=CLR["malicious_fg"])

        # Detection ratio – center-aligned
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = 60

    # Auto-filter on headers
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(results)+2}"

    # ── Sheet 2: Statistics ──────────────────────────────────────────────── #
    ws2 = wb.create_sheet("Statistics")
    ws2.sheet_view.showGridLines = False

    # Compute summary counts
    clean_count      = sum(1 for r in results if not r.get("error") and r.get("malicious",0)==0 and r.get("suspicious",0)==0 and r.get("total_engines",0)>0)
    malicious_count  = sum(1 for r in results if not r.get("error") and r.get("malicious",0)>0)
    suspicious_count = sum(1 for r in results if not r.get("error") and r.get("malicious",0)==0 and r.get("suspicious",0)>0)
    unknown_count    = sum(1 for r in results if r.get("error") or r.get("total_engines",0)==0)

    # Title
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value     = "Summary Statistics"
    t2.font      = Font(name="Arial", size=14, bold=True, color=CLR["header_fg"])
    t2.fill      = clr(CLR["title_bg"])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Stats table
    stats_data = [
        ("Category",              "Count", "Percentage"),
        ("Clean",                 clean_count,      f"=B3/B7*100"),
        ("Malicious",             malicious_count,  f"=B4/B7*100"),
        ("Suspicious",            suspicious_count, f"=B5/B7*100"),
        ("Unknown / Error",       unknown_count,    f"=B6/B7*100"),
        ("TOTAL",                 len(results),     "100%"),
    ]

    cat_fills = [CLR["subheader_bg"], CLR["clean_bg"], CLR["malicious_bg"], CLR["suspicious_bg"], CLR["unknown_bg"], CLR["header_bg"]]
    cat_fgs   = [CLR["header_fg"],    CLR["clean_fg"], CLR["malicious_fg"], CLR["suspicious_fg"], CLR["unknown_fg"], CLR["header_fg"]]

    for r_off, (row, bg, fg) in enumerate(zip(stats_data, cat_fills, cat_fgs), 2):
        for c_off, val in enumerate(row, 1):
            cell           = ws2.cell(row=r_off, column=c_off, value=val)
            cell.fill      = clr(bg)
            cell.font      = Font(name="Arial", size=10, bold=(r_off in (2, 7)), color=fg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
        ws2.row_dimensions[r_off].height = 22

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    # ── Pie chart: verdict distribution ──────────────────────────────────── #
    pie = PieChart()
    pie.title  = "Verdict Distribution"
    pie.style  = 10
    pie.width  = 16
    pie.height = 14

    labels = Reference(ws2, min_col=1, min_row=3, max_row=6)
    data   = Reference(ws2, min_col=2, min_row=2, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True

    # Colour slices
    slice_colors = ["00B050", "FF0000", "FFC000", "808080"]
    for i, hex_c in enumerate(slice_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = hex_c
        pie.series[0].dPt.append(pt)

    ws2.add_chart(pie, "E2")

    # ── Bar chart: detection counts per hash ─────────────────────────────── #
    ws3 = wb.create_sheet("Detection Chart")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    t3 = ws3["A1"]
    t3.value     = "Detection Counts per Hash"
    t3.font      = Font(name="Arial", size=14, bold=True, color=CLR["header_fg"])
    t3.fill      = clr(CLR["title_bg"])
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    for col, hdr in enumerate(["Hash (short)", "Malicious", "Suspicious", "Undetected"], 1):
        c = ws3.cell(row=2, column=col, value=hdr)
        c.font      = Font(name="Arial", size=10, bold=True, color=CLR["header_fg"])
        c.fill      = clr(CLR["header_bg"])
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()

    for r_off, rec in enumerate(results, 3):
        short_hash = (rec.get("sha256") or rec.get("hash",""))[:16] + "…"
        ws3.cell(row=r_off, column=1, value=short_hash)
        ws3.cell(row=r_off, column=2, value=rec.get("malicious",  0) or 0)
        ws3.cell(row=r_off, column=3, value=rec.get("suspicious", 0) or 0)
        ws3.cell(row=r_off, column=4, value=rec.get("undetected", 0) or 0)
        ws3.row_dimensions[r_off].height = 16

    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14

    last_data_row = len(results) + 2
    bar = BarChart()
    bar.type    = "col"
    bar.title   = "Detection Breakdown per Hash"
    bar.y_axis.title = "Engine Count"
    bar.x_axis.title = "File Hash"
    bar.style   = 10
    bar.width   = 30
    bar.height  = 18

    cats  = Reference(ws3, min_col=1, min_row=3, max_row=last_data_row)
    data1 = Reference(ws3, min_col=2, min_row=2, max_row=last_data_row)
    data2 = Reference(ws3, min_col=3, min_row=2, max_row=last_data_row)
    data3 = Reference(ws3, min_col=4, min_row=2, max_row=last_data_row)
    bar.add_data(data1, titles_from_data=True)
    bar.add_data(data2, titles_from_data=True)
    bar.add_data(data3, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "FF0000"
    bar.series[1].graphicalProperties.solidFill = "FFC000"
    bar.series[2].graphicalProperties.solidFill = "00B050"

    ws3.add_chart(bar, "F2")

    # ── Save workbook ─────────────────────────────────────────────────────── #
    wb.save(output_path)
    print(f"\n  ✔ Report saved to: {output_path}")


# ─────────────────────────── MAIN ─────────────────────────────────────────── #

def collect_hashes() -> list:
    """Interactively collect hashes from the user."""
    print("\n" + "="*60)
    print("  VirusTotal Hash Analyzer")
    print("="*60)
    print("Enter file hashes (MD5 / SHA1 / SHA256), one per line.")
    print("Press ENTER on a blank line when done.\n")

    hashes = []
    while True:
        try:
            line = input("  Hash> ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not line:
            if hashes:
                break
            print("  (no hashes entered yet)")
            continue
        # Basic length sanity check
        if len(line) not in (32, 40, 64):
            print(f"  ⚠  '{line[:20]}…' doesn't look like MD5/SHA1/SHA256 – added anyway.")
        hashes.append(line)

    return hashes


if __name__ == "__main__":
    # ── Optionally hard-code hashes here for non-interactive use ──────────── #
    HASHES = []   # e.g. ["44d88612fea8a8f36de82e1278abb02f", ...]

    if HASHES:
        hashes = HASHES
    else:
        hashes = collect_hashes()

    if not hashes:
        print("No hashes provided. Exiting.")
        sys.exit(0)

    print(f"\n  {len(hashes)} hash(es) collected. Starting VirusTotal queries …\n")
    results = query_all(hashes)

    print("\n  Building Excel report …")
    build_excel(results, OUTPUT)
    print("\nDone! Open VirusTotal_Report.xlsx to view results.\n")
